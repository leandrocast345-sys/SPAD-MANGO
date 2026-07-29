import cv2                          # OpenCV: captura y procesamiento de imágenes
import tkinter as tk                 # Tkinter: librería para construir la interfaz gráfica
from tkinter import ttk, messagebox  # ttk: Treeview del historial · messagebox: confirmaciones al eliminar datos
from PIL import Image, ImageTk       # Pillow: convierte imágenes OpenCV a formato que Tkinter puede mostrar
import threading                     # Permite correr el video en un hilo separado sin congelar la UI
import numpy as np                   # NumPy: operaciones matemáticas sobre arrays (promedios, máscaras, etc.)
import time                          # Para pausar cuando falla la conexión
import queue                         # Cola para desacoplar las escrituras al Excel del hilo de análisis/UI
import os                            # Para ubicar el archivo Excel junto al script
from datetime import datetime        # Marca de tiempo de cada sesión guardada
from openpyxl import Workbook, load_workbook   # Lectura/escritura del historial en Excel (.xlsx)
from openpyxl.styles import Font               # Estilo del encabezado del Excel
BG          = "#09110D" ## Color de fondo principal de la ventana
PANEL       = "#141F18" ## Color de fondo de los paneles de video y resultados
BORDER      = "#1E3028" ## Color de borde de los paneles y divisores
ACCENT      = "#4ADE80" ## Color de acento principal (verde) para botones y texto destacado
ACCENT_WARM = "#FCD34D" ## Color de acento secundario (amarillo) para advertencias
TEXT_SEC    = "#6EAA84" ## Color de texto secundario (gris verdoso) para etiquetas y descripciones
TEXT_DIM    = "#2E4A38" ## Color de texto atenuado (gris oscuro) para información menos relevante
BTN_BG      = "#1A3B27" ## Color de fondo de los botones
BTN_HOVER   = "#255C3C" ## Color de fondo de los botones cuando el mouse pasa por encima
RED         = "#F87171" ## Color rojo para alertas y errores
BLUE        = "#60A5FA" ## Color azul para información destacada
FONT_MONO  = ("Courier New", 10) ## Fuente monoespaciada para mostrar valores de SPAD y porcentajes
FONT_BIG   = ("Courier New", 28, "bold") ## Fuente grande y en negrita para mostrar el valor SPAD promedio
FONT_MED   = ("Courier New", 13, "bold") ## Fuente mediana y en negrita para mostrar valores de SPAD individuales
NUM_CAPTURAS = 5   # zonas de la hoja a capturar
     
     #pv     = porcentaje de píxeles verdes  (0-100)
     #pa     = porcentaje de píxeles amarillos (0-100)
     #spad   = valor SPAD estimado (correlación basada en estudios generales con el verde)
     #mask_v = máscara binaria de píxeles verdes
    #mask_a = máscara binaria de píxeles amarillos


def analizar_roi(roi): ## Analiza la región de interés (ROI) de la hoja y calcula los porcentajes de verde y amarillo, así como un valor SPAD estimado.
    hsv = cv2.cvtColor(roi, cv2.COLOR_BGR2HSV) ## Convierte la imagen de BGR a HSV para facilitar la segmentación de colores.
    mask_v = cv2.inRange(hsv, np.array([30, 40, 30]),  np.array([90, 255, 200])) ## Crea una máscara binaria para los píxeles verdes dentro del rango especificado.
    mask_a = cv2.inRange(hsv, np.array([10, 40, 40]),  np.array([30, 255, 255]))## Crea una máscara binaria para los píxeles amarillos dentro del rango especificado.
    av = cv2.countNonZero(mask_v) ## Cuenta el número de píxeles verdes en la máscara.
    aa = cv2.countNonZero(mask_a) ## Cuenta el número de píxeles amarillos en la máscara.
    total = av + aa ## Calcula el total de píxeles detectados (verdes + amarillos).
    if total < 500: ## Si el total de píxeles detectados es menor a 500, se considera que la muestra es insuficiente para un análisis confiable.
        return None
    pv   = (av / total) * 100 ## Calcula el porcentaje de píxeles verdes respecto al total.
    pa   = (aa / total) * 100 ## Calcula el porcentaje de píxeles amarillos respecto al total.
    spad = 10 + (pv * 0.4) ## Estima un valor SPAD basado en el porcentaje de verde, usando una correlación lineal simple (10 + 0.4 * pv).
    return pv, pa, spad, mask_v, mask_a ## Devuelve los resultados del análisis: porcentaje de verde, porcentaje de amarillo, valor SPAD estimado y las máscaras binarias correspondientes.

DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "spad_historial.xlsx") ## Ruta del archivo Excel donde se guarda el historial, ubicado junto al script.
HEADERS_RESUMEN    = ["ID Sesion", "Fecha", "SPAD Promedio", "Verde Promedio %", "Amarillo Promedio %", "Desv. Estandar", "Diagnostico"] ## Encabezados de la hoja "Resumen" (una fila por sesión de 5 zonas).
HEADERS_MEDICIONES = ["ID Sesion", "Zona", "Verde %", "Amarillo %", "SPAD"] ## Encabezados de la hoja "Mediciones" (una fila por cada zona capturada).


class ExcelDB: ## Maneja el guardado del historial de mediciones SPAD en un archivo Excel (.xlsx), sin usar SQL.
    """
    Guarda el historial en un archivo Excel con dos hojas:
      - "Resumen"    -> una fila por sesión (fecha + promedios de las 5 zonas).
      - "Mediciones" -> una fila por cada zona capturada (5 por sesión), enlazada
                        a su sesión mediante la columna "ID Sesion".

    Pensado para NO afectar el rendimiento del programa:
      - Todas las escrituras (guardar / eliminar) se encolan y se procesan en un
        único hilo dedicado; el hilo de análisis y la interfaz nunca esperan a
        que el archivo .xlsx se abra o se guarde en disco.
      - El archivo solo se abre y reescribe UNA vez por sesión completa
        (cuando terminan las 5 zonas), nunca en cada frame de video.
      - Las lecturas (para mostrar el historial) sólo ocurren bajo demanda,
        cuando el usuario abre la ventana de historial.
    """
    def __init__(self, path=DB_PATH):
        self.path = path
        self._lock = threading.Lock()          # protege el archivo entre el hilo de escritura y las lecturas bajo demanda
        self._write_q = queue.Queue()
        self._siguiente_id = 1
        self._inicializar_archivo()
        self._worker = threading.Thread(target=self._procesar_cola, daemon=True)
        self._worker.start()

    def _inicializar_archivo(self): ## Crea el archivo si no existe, o calcula el próximo ID de sesión disponible si ya existe.
        with self._lock:
            if not os.path.exists(self.path):
                self._crear_archivo_vacio()
            else:
                wb = load_workbook(self.path)
                ws = wb["Resumen"]
                ids = [row[0].value for row in ws.iter_rows(min_row=2) if row[0].value is not None]
                self._siguiente_id = (max(ids) + 1) if ids else 1

    def _crear_archivo_vacio(self): ## Crea (o reinicia) el archivo Excel con las dos hojas y sus encabezados.
        wb  = Workbook()
        ws1 = wb.active
        ws1.title = "Resumen"
        ws1.append(HEADERS_RESUMEN)
        ws2 = wb.create_sheet("Mediciones")
        ws2.append(HEADERS_MEDICIONES)
        for ws in (ws1, ws2):
            for cell in ws[1]:
                cell.font = Font(name="Arial", bold=True) ## Encabezados en negrita para que el archivo sea legible al abrirlo directamente.
            ws.freeze_panes = "A2"
        wb.save(self.path)
        self._siguiente_id = 1

    # API pública (no bloqueante, sólo encola la operación) 
    def guardar_sesion(self, capturas, avg_pv, avg_pa, avg_spad, std_spad, diagnostico): ## Encola el guardado de una sesión completa (5 zonas + promedios). No bloquea la UI ni el hilo de análisis.
        self._write_q.put({
            "accion": "guardar",
            "fecha": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "capturas": list(capturas),
            "avg_pv": avg_pv, "avg_pa": avg_pa, "avg_spad": avg_spad,
            "std_spad": std_spad, "diagnostico": diagnostico,
        })

    def eliminar_sesion(self, sesion_id, callback=None): ## Encola la eliminación de UNA sesión puntual (su fila de resumen + sus 5 mediciones).
        self._write_q.put({"accion": "eliminar_sesion", "sesion_id": sesion_id, "callback": callback})

    def eliminar_todo(self, callback=None): ## Encola el borrado de TODO el historial (reinicia el archivo Excel vacío).
        self._write_q.put({"accion": "eliminar_todo", "callback": callback})

    # Procesamiento en hilo dedicado 
    def _procesar_cola(self): ## Hilo único que procesa guardados/eliminaciones en orden, para que nunca haya dos escrituras simultáneas sobre el mismo archivo.
        while True:
            payload = self._write_q.get()
            if payload is None:
                break
            try:
                accion = payload["accion"]
                with self._lock:
                    if accion == "guardar":
                        self._escribir_sesion(payload)
                    elif accion == "eliminar_sesion":
                        self._borrar_sesion(payload["sesion_id"])
                    elif accion == "eliminar_todo":
                        self._crear_archivo_vacio()
            except Exception as e:
                print(f"[Excel] Error procesando '{payload.get('accion')}': {e}")
            finally:
                if payload.get("callback"):
                    payload["callback"]()
                self._write_q.task_done()

    def _escribir_sesion(self, payload): ## Abre el Excel, agrega la fila de resumen y las 5 filas de mediciones, y guarda.
        wb = load_workbook(self.path)
        ws_resumen = wb["Resumen"]
        ws_med     = wb["Mediciones"]
        sesion_id  = self._siguiente_id
        self._siguiente_id += 1
        ws_resumen.append([sesion_id, payload["fecha"], round(payload["avg_spad"], 2),
                            round(payload["avg_pv"], 2), round(payload["avg_pa"], 2),
                            round(payload["std_spad"], 2), payload["diagnostico"]])
        for i, c in enumerate(payload["capturas"]):
            ws_med.append([sesion_id, i + 1, round(c["pv"], 2), round(c["pa"], 2), round(c["spad"], 2)])
        wb.save(self.path)

    def _borrar_sesion(self, sesion_id): ## Elimina, en ambas hojas, todas las filas asociadas al ID de sesión indicado.
        wb = load_workbook(self.path)
        for nombre in ("Resumen", "Mediciones"):
            ws = wb[nombre]
            filas = [r for r in range(ws.max_row, 1, -1) if ws.cell(row=r, column=1).value == sesion_id] ## Se recorre de abajo hacia arriba para que borrar una fila no desordene los índices de las siguientes.
            for r in filas:
                ws.delete_rows(r)
        wb.save(self.path)

    # Lectura bajo demanda 
    def obtener_historial(self): ## Lee el Excel completo y arma la lista de sesiones con sus 5 mediciones cada una. Solo se llama al abrir la ventana de Historial.
        with self._lock:
            wb = load_workbook(self.path, data_only=True)
            resumen    = [tuple(r) for r in wb["Resumen"].iter_rows(min_row=2, values_only=True) if r[0] is not None]
            mediciones = [tuple(r) for r in wb["Mediciones"].iter_rows(min_row=2, values_only=True) if r[0] is not None]
        resultado = []
        for r in resumen:
            sid  = r[0]
            meds = sorted((m for m in mediciones if m[0] == sid), key=lambda m: m[1])
            resultado.append({"resumen": r, "mediciones": meds})
        resultado.sort(key=lambda item: item["resumen"][0], reverse=True) ## Sesiones más recientes primero.
        return resultado

    def cerrar(self): ## Detiene el hilo de escritura de forma ordenada al cerrar la aplicación.
        self._write_q.put(None)
        self._worker.join(timeout=2)


class SPADApp: ## Clase principal de la aplicación SPAD Pro, que maneja la interfaz gráfica y la lógica de captura y análisis de imágenes.
    def __init__(self, window): ## Inicializa la aplicación, configurando la ventana principal, los paneles de video y los botones de control.
        self.window = window
        self.window.title("SPAD Pro — Analizador de Clorofila")
        self.window.configure(bg=BG) ## Configura el color de fondo de la ventana principal.
        self.window.resizable(False, False) ## Evita que la ventana sea redimensionable.

        self.PW, self.PH = 300, 210   # tamaño cada panel de cámara

        # Estado multi-captura
        self.capturas     = []   # lista de dicts {pv, pa, spad}
        self.frame_actual = None
        self.running      = True
        self.analizando   = False

        self.db = ExcelDB()   # historial persistente en Excel (spad_historial.xlsx)

        self._build_ui()
        ## conexion con la cámara IP 
        self.url = "http://192.168.1.1:81/stream"  # URL de la cámara IP (ajustar según la configuración de la cámara)
        threading.Thread(target=self.stream_video, daemon=True).start()
        self.actualizar_vista_previa()

    ## INTERFAZ GRÁFICA 

    def _build_ui(self):
        # Encabezado 
        hdr = tk.Frame(self.window, bg=BG) ## Crea un marco para el encabezado de la aplicación, con un fondo del color definido en BG.
        hdr.pack(fill="x", padx=20, pady=(16, 4)) ## Empaqueta el marco del encabezado en la ventana principal, ocupando todo el ancho disponible y con un padding horizontal de 20 píxeles y vertical de 16 píxeles arriba y 4 píxeles abajo.
        tk.Label(hdr, text="◈  SPAD PRO", bg=BG, fg=ACCENT,
                 font=("Segoe UI", 14, "bold")).pack(side="left")
        tk.Label(hdr, text="Análisis Multi-Zona de Clorofila",
                 bg=BG, fg=TEXT_SEC, font=("Segoe UI", 9)).pack(side="left", padx=10)
        self.lbl_status = tk.Label(hdr, text="● EN VIVO", bg=BG,
                                   fg=TEXT_DIM, font=("Segoe UI", 8, "bold"))
        self.lbl_status.pack(side="right")
        tk.Frame(self.window, bg=BORDER, height=1).pack(fill="x", padx=20, pady=4)

        #  Fila superior: video en vivo + última captura
        top = tk.Frame(self.window, bg=BG) ## Crea un marco para la fila superior de la interfaz, que contendrá los paneles de video en vivo y la última captura realizada.
        top.pack(padx=20, pady=6) ## Empaqueta el marco de la fila superior en la ventana principal, con un padding horizontal de 20 píxeles y vertical de 6 píxeles.

        self.canvas_vid    = self._panel(top, "VISTA EN VIVO") ## Crea un panel de video en vivo, donde se mostrará la transmisión de la cámara IP.
        self.canvas_ultima = self._panel(top, "ÚLTIMA CAPTURA") ## Crea un panel para mostrar la última captura realizada, permitiendo al usuario ver el resultado del análisis de la zona capturada.
        self.canvas_verde  = self._panel(top, "ZONA VERDE") ## Crea un panel para mostrar la zona verde detectada en la última captura, utilizando la máscara binaria correspondiente para resaltar los píxeles verdes.
        self.canvas_amarilla = self._panel(top, "ZONA AMARILLA") ## Crea un panel para mostrar la zona amarilla detectada en la última captura, utilizando la máscara binaria correspondiente para resaltar los píxeles amarillos.

        ## CAPTURAS DE ZONAS DE LA HOJA
        tk.Frame(self.window, bg=BORDER, height=1).pack(fill="x", padx=20, pady=(4,2)) ## Crea una línea divisoria horizontal entre la fila superior de paneles y la sección de capturas de zonas de la hoja, utilizando un marco con un color de borde definido en BORDER y un padding horizontal de 20 píxeles y vertical de 4 píxeles arriba y 2 píxeles abajo.
        strip_hdr = tk.Frame(self.window, bg=BG) ## Crea un marco para el encabezado de la sección de capturas de zonas de la hoja, con un fondo del color definido en BG.
        strip_hdr.pack(fill="x", padx=20)
        tk.Label(strip_hdr, text="CAPTURAS DE ZONAS", bg=BG,
                 fg=TEXT_SEC, font=("Segoe UI", 7, "bold")).pack(side="left")
        self.lbl_n = tk.Label(strip_hdr, text="0 / 5", bg=BG,
                               fg=ACCENT, font=("Courier New", 9, "bold"))
        self.lbl_n.pack(side="right")

        self.strip_frame = tk.Frame(self.window, bg=BG)
        self.strip_frame.pack(padx=20, pady=(2, 6))

        self.thumbs      = []   # tk widgets ## Contenedores de miniaturas de cada zona capturada
        self.thumb_imgs  = []   # PhotoImage refs (evitar GC) ## Referencias a las imágenes de las miniaturas para evitar que el recolector de basura las elimine.
        self.spad_lbls   = []   # etiquetas SPAD por miniatura ### Crea etiquetas para mostrar el valor SPAD correspondiente a cada miniatura de zona capturada.

        TW, TH = 180, 100 ## Tamaño de cada miniatura
        for i in range(NUM_CAPTURAS): ## Itera sobre el número de capturas definidas (5 zonas) para crear los contenedores de miniaturas y las etiquetas SPAD correspondientes.
            col = tk.Frame(self.strip_frame, bg=BORDER, padx=1, pady=1)
            col.pack(side="left", padx=4)
            inner = tk.Frame(col, bg=PANEL)
            inner.pack()
            tk.Label(inner, text=f"ZONA {i+1}", bg=PANEL, fg=TEXT_DIM,
                     font=("Segoe UI", 6, "bold")).pack(pady=(3,0))
            c = tk.Canvas(inner, width=TW, height=TH, bg="#080F0B",
                           highlightthickness=0)
            c.pack(padx=0, pady=(1,0))
            c.create_text(TW//2, TH//2, text=f"{i+1}", fill=TEXT_DIM,
                           font=("Courier New", 18), tags="num")
            lbl = tk.Label(inner, text="—", bg=PANEL, fg=TEXT_DIM,
                           font=("Courier New", 9, "bold"))
            lbl.pack(pady=(1,3))
            self.thumbs.append((c, TW, TH))
            self.thumb_imgs.append(None)
            self.spad_lbls.append(lbl)

        #Botones 
        btn_row = tk.Frame(self.window, bg=BG) ## Crea un marco para contener los botones de control de la aplicación, con un fondo del color definido en BG.
        btn_row.pack(pady=(4, 4)) ## Empaqueta el marco de botones en la ventana principal, con un padding vertical de 4 píxeles arriba y abajo.

        self.btn_cap = self._btn(btn_row, " CAPTURAR ZONA", ## Crea un botón principal para capturar la siguiente zona de la hoja, con el texto "CAPTURAR ZONA" y un color de acento definido en ACCENT.
                                  self.capturar_zona, ACCENT)
        self.btn_cap.pack(side="left", padx=6)

        self.btn_reset = self._btn(btn_row, " REINICIAR", ## Crea un botón secundario para reiniciar la aplicación, borrando todas las capturas de zonas realizadas y restableciendo los valores de SPAD y porcentajes a sus valores iniciales, con el texto "REINICIAR" y un color de texto secundario definido en TEXT_SEC.
                                    self.reiniciar, TEXT_SEC)
        self.btn_reset.pack(side="left", padx=6) ## Empaqueta el botón de reinicio a la izquierda del marco de botones, con un padding horizontal de 6 píxeles entre los botones.

        self.btn_hist = self._btn(btn_row, " HISTORIAL", ## Botón para abrir la ventana con el historial de sesiones guardadas en el Excel.
                                   self.mostrar_historial, BLUE)
        self.btn_hist.pack(side="left", padx=6)

        # Panel de resultados promedio 
        tk.Frame(self.window, bg=BORDER, height=1).pack(fill="x", padx=20, pady=(4,0)) ## Crea una línea divisoria horizontal entre la sección de capturas de zonas y el panel de resultados promedio, utilizando un marco con un color de borde definido en BORDER y un padding horizontal de 20 píxeles y vertical de 4 píxeles arriba y 0 píxeles abajo.
        res_outer = tk.Frame(self.window, bg=BORDER)
        res_outer.pack(padx=20, pady=(0,4), fill="x")
        res = tk.Frame(res_outer, bg=PANEL, padx=20, pady=12)
        res.pack(fill="x", padx=1, pady=1)

        spad_col = tk.Frame(res, bg=PANEL) ## Crea un marco para la columna de resultados promedio, que mostrará el valor SPAD promedio calculado a partir de las capturas de zonas.
        spad_col.pack(side="left", padx=(0,30))
        tk.Label(spad_col, text="SPAD PROMEDIO", bg=PANEL, fg=TEXT_SEC,
                 font=("Segoe UI", 8, "bold")).pack(anchor="w")
        self.lbl_spad = tk.Label(spad_col, text="—", bg=PANEL,
                                  fg=TEXT_DIM, font=FONT_BIG)
        self.lbl_spad.pack(anchor="w") ## Empaqueta la etiqueta que mostrará el valor SPAD promedio en la columna de resultados, alineándola a la izquierda del marco.
        tk.Label(spad_col, text="unidades relativas  (5 zonas)", bg=PANEL,
                 fg=TEXT_DIM, font=("Segoe UI", 8)).pack(anchor="w")

        tk.Frame(res, bg=BORDER, width=1).pack(side="left", fill="y", padx=10)

        met_col = tk.Frame(res, bg=PANEL) ## Crea un marco para la columna de métricas, que mostrará los porcentajes promedio de verde y amarillo, así como el diagnóstico y la desviación estándar entre las zonas capturadas.
        met_col.pack(side="left") ## Empaqueta la columna de métricas a la izquierda del marco de resultados, permitiendo que se muestre junto a la columna de SPAD promedio.
        self.lbl_verde    = self._metric(met_col, "Verde promedio",    "—%",  ACCENT)
        self.lbl_amarillo = self._metric(met_col, "Amarillo promedio", "—%",  ACCENT_WARM)
        self.lbl_diag     = self._metric(met_col, "Diagnóstico",       "Captura 5 zonas de la hoja", TEXT_SEC)
        self.lbl_std      = self._metric(met_col, "Desv. estándar",    "—",   TEXT_SEC)

        # Barra cromática
        bar_frame = tk.Frame(self.window, bg=BG) ## Crea un marco para contener la barra cromática que mostrará la distribución promedio de los colores verde y amarillo en las zonas capturadas.   
        bar_frame.pack(padx=20, pady=(2,4), fill="x")
        tk.Label(bar_frame, text="DISTRIBUCIÓN CROMÁTICA PROMEDIO", bg=BG,
                 fg=TEXT_SEC, font=("Segoe UI", 7, "bold")).pack(anchor="w")
        bar_outer = tk.Frame(bar_frame, bg=BORDER, height=10)
        bar_outer.pack(fill="x", pady=3)
        bar_outer.pack_propagate(False)
        self.bar_canvas = tk.Canvas(bar_outer, height=10, bg=TEXT_DIM,
                                     highlightthickness=0)
        self.bar_canvas.pack(fill="both")

        # Pie
        tk.Frame(self.window, bg=BORDER, height=1).pack(fill="x", padx=20, pady=(4,0)) ## Crea una línea divisoria horizontal entre el panel de resultados y el pie de la aplicación, utilizando un marco con un color de borde definido en BORDER y un padding horizontal de 20 píxeles y vertical de 4 píxeles arriba y 0 píxeles abajo.
        footer = tk.Frame(self.window, bg=BG)
        footer.pack(fill="x", padx=20, pady=5)
        tk.Label(footer,
                 text="Captura 5 zonas distintas de la hoja → el sistema promedia y muestra el SPAD final  ·  Calibrar con SPAD-502 para valores absolutos",
                 bg=BG, fg=TEXT_DIM, font=("Segoe UI", 8)).pack(side="left")

    def _panel(self, parent, titulo): ## Crea un panel de video con un título, que se utilizará para mostrar la transmisión en vivo de la cámara IP, la última captura realizada y las zonas verde y amarilla detectadas en la hoja.
        wrapper = tk.Frame(parent, bg=BORDER, padx=1, pady=1)
        wrapper.pack(side="left", padx=5)
        inner = tk.Frame(wrapper, bg=PANEL)
        inner.pack()
        tk.Label(inner, text=titulo, bg=PANEL, fg=TEXT_SEC,
                 font=("Segoe UI", 7, "bold")).pack(anchor="w", padx=5, pady=(3,0))
        c = tk.Canvas(inner, width=self.PW, height=self.PH,
                       bg="#080F0B", highlightthickness=0)
        c.pack(padx=0, pady=(2,4))
        c.create_text(self.PW//2, self.PH//2, text="SIN SEÑAL",
                       fill=TEXT_DIM, font=FONT_MONO, tags="nosig")
        return c

    def _btn(self, parent, texto, cmd, color): ## Crea un botón con estilo personalizado, que se utilizará para los controles de la aplicación, como capturar zonas y reiniciar el análisis.
        b = tk.Button(parent, text=texto, command=cmd,
                      bg=BTN_BG, fg=color,
                      activebackground=BTN_HOVER, activeforeground=color,
                      font=("Segoe UI", 11, "bold"),
                      relief="flat", bd=0, padx=24, pady=8, cursor="hand2")
        b.bind("<Enter>", lambda e: b.config(bg=BTN_HOVER))
        b.bind("<Leave>", lambda e: b.config(bg=BTN_BG))
        return b

    def _metric(self, parent, label, valor, color): ## Crea un widget de métrica que muestra una etiqueta y un valor, utilizado para mostrar los porcentajes promedio de verde y amarillo, el diagnóstico y la desviación estándar entre las zonas capturadas.
        row = tk.Frame(parent, bg=PANEL)
        row.pack(anchor="w", pady=1)
        tk.Label(row, text=f"{label}:", bg=PANEL, fg=TEXT_SEC,
                 font=("Segoe UI", 8), width=20, anchor="w").pack(side="left")
        lbl = tk.Label(row, text=valor, bg=PANEL, fg=color,
                        font=("Segoe UI", 9, "bold"))
        lbl.pack(side="left", padx=4)
        return lbl
    # Video
    def stream_video(self): ## Inicia un hilo para capturar el video en vivo desde la cámara IP, actualizando el estado de la conexión y almacenando el último frame capturado para su posterior análisis y visualización en la interfaz gráfica.
        cap = cv2.VideoCapture(self.url)
        while self.running:
            ret, frame = cap.read()
            if ret:
                self.frame_actual = frame
                self.lbl_status.config(fg=ACCENT)
            else:
                self.lbl_status.config(fg=RED)
                time.sleep(0.5)
        cap.release()

    def _mostrar(self, img_bgr, canvas, overlay=None): ## Muestra una imagen en un canvas de Tkinter, convirtiendo la imagen de BGR a RGB y redimensionándola al tamaño del panel correspondiente. Si se proporciona un texto de superposición, se dibuja sobre la imagen.
        img_rgb = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2RGB)
        img_pil = Image.fromarray(img_rgb).resize((self.PW, self.PH))
        img_tk  = ImageTk.PhotoImage(img_pil)
        canvas.delete("nosig")
        canvas.create_image(0, 0, anchor="nw", image=img_tk)
        canvas.image = img_tk
        if overlay:
            canvas.create_text(6, self.PH - 14, anchor="w",
                                text=overlay, fill=ACCENT, font=FONT_MONO)

    def _mostrar_thumb(self, idx, img_bgr, spad): ## Muestra una miniatura de la zona capturada en el panel correspondiente, redimensionando la imagen al tamaño de la miniatura y actualizando la etiqueta SPAD asociada.
        c, TW, TH = self.thumbs[idx]
        img_rgb = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2RGB)
        img_pil = Image.fromarray(img_rgb).resize((TW, TH))
        img_tk  = ImageTk.PhotoImage(img_pil)
        c.delete("num")
        c.create_image(0, 0, anchor="nw", image=img_tk)
        c.image = img_tk
        self.thumb_imgs[idx] = img_tk
        self.spad_lbls[idx].config(text=f"SPAD {spad:.1f}", fg=ACCENT)

    def actualizar_vista_previa(self): ### Actualiza la vista previa en vivo del video de la cámara IP, dibujando un recuadro de análisis en el centro del panel y mostrando el número de zona actual que se está capturando. Esta función se llama periódicamente para mantener la vista previa actualizada.
        if self.frame_actual is not None:
            h, w = self.frame_actual.shape[:2]
            x1, y1 = w//10, h//10
            x2, y2 = 9*w//10, 9*h//10
            fg = self.frame_actual.copy()
            cv2.rectangle(fg, (x1,y1), (x2,y2), (80,160,255), 2) ## Dibuja un recuadro azul en el centro del panel de video en vivo, indicando la región de interés (ROI) que se analizará al capturar una zona de la hoja.
            L = 16
            for px,py,dx,dy in [(x1,y1,1,1),(x2,y1,-1,1),(x1,y2,1,-1),(x2,y2,-1,-1)]: ## Dibuja líneas diagonales en las esquinas del recuadro de análisis para resaltar visualmente la región de interés (ROI) que se analizará al capturar una zona de la hoja.
                cv2.line(fg,(px,py),(px+dx*L,py),(80,200,255),3) ## Dibuja una línea horizontal en la esquina superior izquierda del recuadro de análisis.
                cv2.line(fg,(px,py),(px,py+dy*L),(80,200,255),3)## Dibuja una línea vertical en la esquina superior izquierda del recuadro de análisis.
            # Contador en vivo
            n = len(self.capturas) ## Obtiene el número de zonas capturadas hasta el momento, para mostrarlo en la vista previa del video en vivo.
            cv2.putText(fg, f"Zona {n+1}/5", (x1+6, y1+22), ## Dibuja el texto que indica el número de zona actual que se está capturando en la esquina superior izquierda del recuadro de análisis, utilizando la fuente Hershey Simplex y un color azul claro.
                        cv2.FONT_HERSHEY_SIMPLEX, 0.55, (80,220,255), 1, cv2.LINE_AA) ## Dibuja el texto que indica el número de zona actual que se está capturando en la esquina superior izquierda del recuadro de análisis, utilizando la fuente Hershey Simplex y un color azul claro.
            self._mostrar(fg, self.canvas_vid) ## Muestra la imagen con el recuadro de análisis y el número de zona actual en el panel de video en vivo.
        self.window.after(30, self.actualizar_vista_previa) ## Programa la siguiente actualización de la vista previa en 30 milisegundos, para mantener la vista previa del video en vivo actualizada de manera continua.
    
    # Captura y análisis

    def capturar_zona(self): ## Captura la zona de la hoja que se encuentra dentro del recuadro de análisis en el panel de video en vivo, analiza la región de interés (ROI) para calcular los porcentajes de verde y amarillo, así como un valor SPAD estimado, y actualiza la interfaz gráfica con los resultados obtenidos.
        if self.frame_actual is None or self.analizando:
            return
        if len(self.capturas) >= NUM_CAPTURAS:
            return   # ya tenemos todas

        self.analizando = True
        self.btn_cap.config(state="disabled", text="Procesando…", fg=TEXT_DIM)

        def _run(): ## Función interna que se ejecuta en un hilo separado para procesar la captura de la zona de la hoja, evitando que la interfaz gráfica se congele durante el análisis.
            frame = self.frame_actual.copy()
            h, w  = frame.shape[:2] ## Altura y ancho del frame
            x1, y1 = w//10, h//10 ## Coordenadas del recuadro de análisis (10% a 90% del ancho y alto)
            x2, y2 = 9*w//10, 9*h//10 ## Coordenadas del recuadro de análisis (10% a 90% del ancho y alto)
            roi = frame[y1:y2, x1:x2] ## Recorte del frame para analizar solo la zona central
            res = analizar_roi(roi)
            if res is None: 
                self.lbl_diag.config(
                    text="Muestra insuficiente — reencuadra la hoja", fg=RED)
                self.btn_cap.config(state="normal",
                                    text=" CAPTURAR ZONA", fg=ACCENT)
                self.analizando = False
                return

            pv, pa, spad, mask_v, mask_a = res
            idx = len(self.capturas)
            self.capturas.append({"pv": pv, "pa": pa, "spad": spad})

            # Mostrar última captura y sus máscaras
            self._mostrar(roi, self.canvas_ultima, overlay=f"Zona {idx+1} — SPAD {spad:.1f}")
            self._mostrar(cv2.bitwise_and(roi, roi, mask=mask_v),
                           self.canvas_verde,
                           overlay=f"Verde {pv:.1f}%")
            self._mostrar(cv2.bitwise_and(roi, roi, mask=mask_a),
                           self.canvas_amarilla,
                           overlay=f"Amarillo {pa:.1f}%")
            self._mostrar_thumb(idx, roi, spad)
            self.lbl_n.config(text=f"{idx+1} / {NUM_CAPTURAS}")

            if len(self.capturas) == NUM_CAPTURAS:
                self._calcular_promedio()
                self.btn_cap.config(state="disabled",
                                    text="5 zonas completadas", fg=TEXT_SEC)
            else:
                self.btn_cap.config(state="normal",
                                    text=f"CAPTURAR ZONA  ({idx+1}/{NUM_CAPTURAS})",
                                    fg=ACCENT)
            self.analizando = False

        threading.Thread(target=_run, daemon=True).start()

    def _calcular_promedio(self): ## Calcula los promedios de los porcentajes de verde y amarillo, así como el valor SPAD promedio y la desviación estándar entre las zonas capturadas, actualizando la interfaz gráfica con los resultados obtenidos y proporcionando un diagnóstico basado en el valor SPAD promedio.
        pvs   = [c["pv"]   for c in self.capturas]
        pas   = [c["pa"]   for c in self.capturas]
        spads = [c["spad"] for c in self.capturas]

        avg_pv   = np.mean(pvs) # Calcula el promedio de los porcentajes de verde de las zonas capturadas.
        avg_pa   = np.mean(pas) # Calcula el promedio de los porcentajes de amarillo de las zonas capturadas.
        avg_spad = np.mean(spads)# Calcula el promedio de los valores SPAD de las zonas capturadas.
        std_spad = np.std(spads) # Calcula la desviación estándar de los valores SPAD de las zonas capturadas, lo que indica la variabilidad entre las zonas.

        if avg_spad >= 45: # Clorofila óptima
            diag, col = "Clorofila óptima ", ACCENT
        elif avg_spad >= 30: ## Clorofila moderada
            diag, col = "Nivel moderado — revisar", ACCENT_WARM
        else:
            diag, col = "Clorosis detectada ", RED

        self.lbl_spad.config(text=f"{avg_spad:.1f}", fg=ACCENT)
        self.lbl_verde.config(text=f"{avg_pv:.1f}%")
        self.lbl_amarillo.config(text=f"{avg_pa:.1f}%")
        self.lbl_diag.config(text=diag, fg=col)
        self.lbl_std.config(text=f"± {std_spad:.2f}  (variación entre zonas)")

        self._actualizar_barra(avg_pv, avg_pa)

        # Guarda la sesión (5 mediciones + promedio) en el Excel. No bloquea: solo se encola.
        self.db.guardar_sesion(self.capturas, avg_pv, avg_pa, avg_spad, std_spad, diag)

    def _actualizar_barra(self, pv, pa): ## Actualiza la barra cromática que muestra la distribución promedio de los colores verde y amarillo en las zonas capturadas, dibujando rectángulos proporcionales a los porcentajes de verde y amarillo calculados.
        self.bar_canvas.update_idletasks()
        W = self.bar_canvas.winfo_width()
        H = 10
        self.bar_canvas.delete("all")
        self.bar_canvas.create_rectangle(0, 0, W, H, fill=TEXT_DIM, outline="")
        xv = int(W * pv / 100) ## Calcula la posición final del rectángulo verde en la barra cromática, proporcional al porcentaje de verde promedio.
        self.bar_canvas.create_rectangle(0, 0, xv, H, fill=ACCENT, outline="")
        xa = int(W * pa / 100) ## Calcula la posición final del rectángulo amarillo en la barra cromática, proporcional al porcentaje de amarillo promedio.
        self.bar_canvas.create_rectangle(xv, 0, xv+xa, H, fill=ACCENT_WARM, outline="")

    def reiniciar(self): ## Reinicia la aplicación, borrando todas las capturas de zonas realizadas, restableciendo los valores de SPAD y porcentajes a sus valores iniciales y limpiando los paneles de video y miniaturas.
        self.capturas = [] ## Reinicia la lista de capturas, eliminando todas las zonas capturadas previamente.
        self.thumb_imgs = [None] * NUM_CAPTURAS
        for i, (c, TW, TH) in enumerate(self.thumbs):
            c.delete("all")
            c.create_text(TW//2, TH//2, text=f"{i+1}", fill=TEXT_DIM,
                           font=("Courier New", 18), tags="num")
            self.spad_lbls[i].config(text="—", fg=TEXT_DIM)

        self.lbl_n.config(text="0 / 5")
        self.lbl_spad.config(text="—", fg=TEXT_DIM)
        self.lbl_verde.config(text="—%")
        self.lbl_amarillo.config(text="—%")
        self.lbl_diag.config(text="Captura 5 zonas de la hoja", fg=TEXT_SEC)
        self.lbl_std.config(text="—")
        self.bar_canvas.delete("all")

        # Limpiar paneles
        for canvas in [self.canvas_ultima, self.canvas_verde, self.canvas_amarilla]: ## Itera sobre los paneles de video que muestran la última captura, la zona verde y la zona amarilla, borrando cualquier contenido previo y mostrando un mensaje de "SIN SEÑAL" para indicar que no hay datos disponibles.
            canvas.delete("all")
            canvas.create_text(self.PW//2, self.PH//2, text="SIN SEÑAL", ## Dibuja el texto "SIN SEÑAL" en el centro del panel, utilizando la fuente monoespaciada definida en FONT_MONO y un color de texto atenuado definido en TEXT_DIM.
                                fill=TEXT_DIM, font=FONT_MONO, tags="nosig") ## Dibuja el texto "SIN SEÑAL" en el centro del panel, utilizando la fuente monoespaciada definida en FONT_MONO y un color de texto atenuado definido en TEXT_DIM.

        self.btn_cap.config(state="normal", ## Habilita el botón de captura de zona, permitiendo al usuario iniciar una nueva captura después de reiniciar la aplicación.
                             text=" CAPTURAR ZONA", fg=ACCENT) ## Cambia el texto del botón de captura de zona a "CAPTURAR ZONA" y establece el color del texto en el color de acento definido en ACCENT.

    # Historial (Excel)

    def mostrar_historial(self): ## Abre una ventana con el historial de sesiones guardadas en el Excel: cada sesión muestra fecha + promedio, y se puede desplegar para ver sus 5 mediciones. Incluye opciones para eliminar una sesión puntual o borrar todo el historial.
        win = tk.Toplevel(self.window)
        win.title("Historial de Mediciones — Excel")
        win.configure(bg=BG)
        win.geometry("700x480")

        tk.Label(win, text="◈ HISTORIAL DE SESIONES", bg=BG, fg=ACCENT,
                 font=("Segoe UI", 12, "bold")).pack(anchor="w", padx=14, pady=(14, 2))
        tk.Label(win, text=f"Archivo: {os.path.basename(self.db.path)}", bg=BG, fg=TEXT_DIM,
                 font=("Segoe UI", 8)).pack(anchor="w", padx=14, pady=(0, 8))

        style = ttk.Style(win) ## Estiliza el Treeview para que combine con la paleta oscura del resto de la app.
        style.theme_use("clam")
        style.configure("Treeview", background=PANEL, fieldbackground=PANEL,
                        foreground="white", rowheight=24, font=("Courier New", 9), borderwidth=0)
        style.configure("Treeview.Heading", background=BTN_BG, foreground=ACCENT,
                        font=("Segoe UI", 9, "bold"))
        style.map("Treeview", background=[("selected", BTN_HOVER)])

        tree_frame = tk.Frame(win, bg=BG)
        tree_frame.pack(fill="both", expand=True, padx=14)

        cols = ("verde", "amarillo", "spad")
        tree = ttk.Treeview(tree_frame, columns=cols, show="tree headings")
        tree.heading("#0", text="Sesión / Zona")
        tree.heading("verde", text="Verde %")
        tree.heading("amarillo", text="Amarillo %")
        tree.heading("spad", text="SPAD")
        tree.column("#0", width=280)
        tree.column("verde", width=90, anchor="center")
        tree.column("amarillo", width=90, anchor="center")
        tree.column("spad", width=130, anchor="center")
        tree.pack(side="left", fill="both", expand=True)

        scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
        scrollbar.pack(side="right", fill="y")
        tree.configure(yscroll=scrollbar.set)

        lbl_estado = tk.Label(win, text="Cargando historial…", bg=BG, fg=TEXT_DIM, font=("Segoe UI", 8))
        lbl_estado.pack(anchor="w", padx=14, pady=(4, 0))

        sesion_por_iid = {}   # iid del árbol -> ID de sesión en el Excel (solo filas raíz = sesiones)

        def poblar(historial): ## Vuelca la lista de sesiones en el Treeview (se ejecuta en el hilo de la UI).
            tree.delete(*tree.get_children())
            sesion_por_iid.clear()
            if not historial:
                lbl_estado.config(text="Todavía no hay sesiones guardadas.")
                return
            for item in historial:
                sid, fecha, avg_spad, avg_pv, avg_pa, std_spad, diag = item["resumen"]
                iid = tree.insert("", "end", text=f"{fecha}  —  {diag}",
                                   values=(f"{avg_pv:.1f}%", f"{avg_pa:.1f}%",
                                           f"{avg_spad:.1f} (± {std_spad:.2f})"))
                sesion_por_iid[iid] = sid
                for zona, pv, pa, spad in item["mediciones"]:
                    tree.insert(iid, "end", text=f"   Zona {zona}",
                                values=(f"{pv:.1f}%", f"{pa:.1f}%", f"{spad:.1f}"))
            lbl_estado.config(text=f"{len(historial)} sesión(es) guardada(s)")

        def cargar(): ## Lee el Excel en un hilo aparte (por si el archivo ya es grande) y actualiza la UI cuando termina.
            historial = self.db.obtener_historial()
            self.window.after(0, lambda: poblar(historial))

        threading.Thread(target=cargar, daemon=True).start()

        # Botones para eliminar datos 
        btns = tk.Frame(win, bg=BG)
        btns.pack(fill="x", padx=14, pady=12)

        def eliminar_seleccion(): ## Elimina del Excel las sesiones seleccionadas en el árbol (ignora si lo seleccionado es una zona individual, no una sesión).
            raiz_sel = [iid for iid in tree.selection() if iid in sesion_por_iid]
            if not raiz_sel:
                messagebox.showinfo("Eliminar", "Selecciona una sesión (no una zona individual).", parent=win)
                return
            if not messagebox.askyesno("Eliminar sesión",
                                        f"¿Eliminar {len(raiz_sel)} sesión(es) del Excel? Esta acción no se puede deshacer.",
                                        parent=win):
                return
            for iid in raiz_sel:
                self.db.eliminar_sesion(sesion_por_iid[iid])
                tree.delete(iid)
                del sesion_por_iid[iid]
            lbl_estado.config(text="Sesión(es) eliminada(s).")

        def borrar_todo(): ## Borra por completo el historial guardado en el Excel (reinicia el archivo).
            if not messagebox.askyesno("Borrar historial completo",
                                        "Esto eliminará TODAS las sesiones guardadas en el Excel. ¿Continuar?",
                                        parent=win):
                return
            self.db.eliminar_todo()
            tree.delete(*tree.get_children())
            sesion_por_iid.clear()
            lbl_estado.config(text="Historial borrado.")

        self._btn(btns, " Eliminar sesión seleccionada", eliminar_seleccion, RED).pack(side="left", padx=(0, 8))
        self._btn(btns, " Borrar todo el historial", borrar_todo, RED).pack(side="left")

    def on_closing(self): ## Maneja el evento de cierre de la ventana principal, deteniendo la captura de video, cerrando el Excel de forma ordenada y destruyendo la ventana para finalizar la aplicación.
        self.running = False
        self.db.cerrar()
        self.window.destroy()


# Arranque 
root = tk.Tk() # Crea la ventana principal de la aplicación utilizando Tkinter.
app  = SPADApp(root) # Inicializa la aplicación SPADApp, pasando la ventana principal como argumento para configurar la interfaz gráfica y la lógica de captura y análisis de imágenes.
root.protocol("WM_DELETE_WINDOW", app.on_closing) # Configura el protocolo de cierre de la ventana principal, asignando la función on_closing de la aplicación para manejar el evento de cierre y detener la captura de video de manera ordenada.
root.mainloop() # Inicia el bucle principal de la interfaz gráfica, permitiendo que la aplicación responda a eventos y actualice la ventana principal de manera continua hasta que se cierre.