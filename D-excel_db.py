"""
excel_db.py
===========
Persistencia del historial de mediciones SPAD en un archivo Excel (.xlsx),
sin usar SQL. Contiene la clase ExcelDB, reutilizable en cualquier proyecto
que necesite guardar/leer sesiones de mediciones de forma asíncrona.
"""

import os
import threading
import queue
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from config import DB_PATH, HEADERS_RESUMEN, HEADERS_MEDICIONES


class ExcelDB:
    """
    Maneja el guardado del historial de mediciones SPAD en un archivo Excel
    (.xlsx), sin usar SQL.

    Guarda el historial en un archivo Excel con dos hojas:
      - "Resumen"    -> una fila por sesión (fecha + promedios de las 5 zonas).
      - "Mediciones" -> una fila por cada zona capturada (5 por sesión), enlazada
                        a su sesión mediante la columna "ID Sesion".

    Pensado para NO afectar el rendimiento del programa:
      - Todas las escrituras (guardar / eliminar) se encolan y se procesan en un
        único hilo dedicado; el hilo de análisis y la interfaz nunca esperan a
        que el archivo .xlsx se abra o se guarde en disco.
      - El archivo solo se abre y reescribe UNA vez por sesión completa
        (cuando terminan las 5 zonas), nunca en cada frame de video.
      - Las lecturas (para mostrar el historial) sólo ocurren bajo demanda,
        cuando el usuario abre la ventana de historial.
    """

    def __init__(self, path=DB_PATH):
        self.path = path
        self._lock = threading.Lock()          # protege el archivo entre el hilo de escritura y las lecturas bajo demanda
        self._write_q = queue.Queue()
        self._siguiente_id = 1
        self._inicializar_archivo()
        self._worker = threading.Thread(target=self._procesar_cola, daemon=True)
        self._worker.start()

    def _inicializar_archivo(self):
        """Crea el archivo si no existe, o calcula el próximo ID de sesión disponible si ya existe."""
        with self._lock:
            if not os.path.exists(self.path):
                self._crear_archivo_vacio()
            else:
                wb = load_workbook(self.path)
                ws = wb["Resumen"]
                ids = [row[0].value for row in ws.iter_rows(min_row=2) if row[0].value is not None]
                self._siguiente_id = (max(ids) + 1) if ids else 1

    def _crear_archivo_vacio(self):
        """Crea (o reinicia) el archivo Excel con las dos hojas y sus encabezados."""
        wb  = Workbook()
        ws1 = wb.active
        ws1.title = "Resumen"
        ws1.append(HEADERS_RESUMEN)
        ws2 = wb.create_sheet("Mediciones")
        ws2.append(HEADERS_MEDICIONES)
        for ws in (ws1, ws2):
            for cell in ws[1]:
                cell.font = Font(name="Arial", bold=True)  # Encabezados en negrita para que el archivo sea legible al abrirlo directamente.
            ws.freeze_panes = "A2"
        wb.save(self.path)
        self._siguiente_id = 1

    # API pública (no bloqueante, sólo encola la operación)
    def guardar_sesion(self, capturas, avg_pv, avg_pa, avg_spad, std_spad, diagnostico):
        """Encola el guardado de una sesión completa (5 zonas + promedios). No bloquea la UI ni el hilo de análisis."""
        self._write_q.put({
            "accion": "guardar",
            "fecha": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "capturas": list(capturas),
            "avg_pv": avg_pv, "avg_pa": avg_pa, "avg_spad": avg_spad,
            "std_spad": std_spad, "diagnostico": diagnostico,
        })

    def eliminar_sesion(self, sesion_id, callback=None):
        """Encola la eliminación de UNA sesión puntual (su fila de resumen + sus 5 mediciones)."""
        self._write_q.put({"accion": "eliminar_sesion", "sesion_id": sesion_id, "callback": callback})

    def eliminar_todo(self, callback=None):
        """Encola el borrado de TODO el historial (reinicia el archivo Excel vacío)."""
        self._write_q.put({"accion": "eliminar_todo", "callback": callback})

    # Procesamiento en hilo dedicado
    def _procesar_cola(self):
        """Hilo único que procesa guardados/eliminaciones en orden, para que nunca haya dos escrituras simultáneas sobre el mismo archivo."""
        while True:
            payload = self._write_q.get()
            if payload is None:
                break
            try:
                accion = payload["accion"]
                with self._lock:
                    if accion == "guardar":
                        self._escribir_sesion(payload)
                    elif accion == "eliminar_sesion":
                        self._borrar_sesion(payload["sesion_id"])
                    elif accion == "eliminar_todo":
                        self._crear_archivo_vacio()
            except Exception as e:
                print(f"[Excel] Error procesando '{payload.get('accion')}': {e}")
            finally:
                if payload.get("callback"):
                    payload["callback"]()
                self._write_q.task_done()

    def _escribir_sesion(self, payload):
        """Abre el Excel, agrega la fila de resumen y las 5 filas de mediciones, y guarda."""
        wb = load_workbook(self.path)
        ws_resumen = wb["Resumen"]
        ws_med     = wb["Mediciones"]
        sesion_id  = self._siguiente_id
        self._siguiente_id += 1
        ws_resumen.append([sesion_id, payload["fecha"], round(payload["avg_spad"], 2),
                            round(payload["avg_pv"], 2), round(payload["avg_pa"], 2),
                            round(payload["std_spad"], 2), payload["diagnostico"]])
        for i, c in enumerate(payload["capturas"]):
            ws_med.append([sesion_id, i + 1, round(c["pv"], 2), round(c["pa"], 2), round(c["spad"], 2)])
        wb.save(self.path)

    def _borrar_sesion(self, sesion_id):
        """Elimina, en ambas hojas, todas las filas asociadas al ID de sesión indicado."""
        wb = load_workbook(self.path)
        for nombre in ("Resumen", "Mediciones"):
            ws = wb[nombre]
            filas = [r for r in range(ws.max_row, 1, -1) if ws.cell(row=r, column=1).value == sesion_id]  # Se recorre de abajo hacia arriba para que borrar una fila no desordene los índices de las siguientes.
            for r in filas:
                ws.delete_rows(r)
        wb.save(self.path)

    # Lectura bajo demanda
    def obtener_historial(self):
        """Lee el Excel completo y arma la lista de sesiones con sus 5 mediciones cada una. Solo se llama al abrir la ventana de Historial."""
        with self._lock:
            wb = load_workbook(self.path, data_only=True)
            resumen    = [tuple(r) for r in wb["Resumen"].iter_rows(min_row=2, values_only=True) if r[0] is not None]
            mediciones = [tuple(r) for r in wb["Mediciones"].iter_rows(min_row=2, values_only=True) if r[0] is not None]
        resultado = []
        for r in resumen:
            sid  = r[0]
            meds = sorted((m for m in mediciones if m[0] == sid), key=lambda m: m[1])
            resultado.append({"resumen": r, "mediciones": meds})
        resultado.sort(key=lambda item: item["resumen"][0], reverse=True)  # Sesiones más recientes primero.
        return resultado

    def cerrar(self):
        """Detiene el hilo de escritura de forma ordenada al cerrar la aplicación."""
        self._write_q.put(None)
        self._worker.join(timeout=2)
